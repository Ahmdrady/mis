from __future__ import annotations

"""
Excel Export Module adapted from Merch Intelligence Hub exporter.
Provides styled multi-sheet workbooks with charts using openpyxl.
"""

import io
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExportColors:
    PRIMARY_BLUE = "1F77B4"
    PRIMARY_GREEN = "2CA02C"
    PRIMARY_RED = "D62728"
    PRIMARY_ORANGE = "FF7F0E"

    HEADER_BG = "2C3E50"
    SUBHEADER_BG = "34495E"
    TABLE_HEADER_BG = "5DADE2"
    ALT_ROW_BG = "F8F9F9"
    TOTAL_ROW_BG = "D5DBDB"

    HEADER_TEXT = "FFFFFF"
    NORMAL_TEXT = "000000"
    MUTED_TEXT = "7F8C8D"

    COVER_BG = "1A5490"
    SECTION_BORDER = "34495E"

    POSITIVE_BG = "D5F4E6"
    NEGATIVE_BG = "FADBD8"
    NEUTRAL_BG = "F8F9F9"


class ExportStyles:
    @staticmethod
    def cover_title():
        return {
            "font": Font(name="Arial", size=24, bold=True, color=ExportColors.HEADER_TEXT),
            "fill": PatternFill("solid", fgColor=ExportColors.COVER_BG),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

    @staticmethod
    def cover_subtitle():
        return {
            "font": Font(name="Arial", size=14, italic=True, color=ExportColors.HEADER_TEXT),
            "fill": PatternFill("solid", fgColor=ExportColors.COVER_BG),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

    @staticmethod
    def section_header():
        return {
            "font": Font(name="Arial", size=12, bold=True, color=ExportColors.HEADER_TEXT),
            "fill": PatternFill("solid", fgColor=ExportColors.HEADER_BG),
            "alignment": Alignment(horizontal="left", vertical="center"),
        }

    @staticmethod
    def table_header():
        return {
            "font": Font(name="Arial", size=11, bold=True, color=ExportColors.HEADER_TEXT),
            "fill": PatternFill("solid", fgColor=ExportColors.TABLE_HEADER_BG),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin", color=ExportColors.SECTION_BORDER),
                right=Side(style="thin", color=ExportColors.SECTION_BORDER),
                top=Side(style="thin", color=ExportColors.SECTION_BORDER),
                bottom=Side(style="thin", color=ExportColors.SECTION_BORDER),
            ),
        }

    @staticmethod
    def data_cell(is_alt: bool = False):
        bg = ExportColors.ALT_ROW_BG if is_alt else "FFFFFF"
        return {
            "font": Font(name="Arial", size=10, color=ExportColors.NORMAL_TEXT),
            "fill": PatternFill("solid", fgColor=bg),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin", color=ExportColors.ALT_ROW_BG),
                right=Side(style="thin", color=ExportColors.ALT_ROW_BG),
                top=Side(style="thin", color=ExportColors.ALT_ROW_BG),
                bottom=Side(style="thin", color=ExportColors.ALT_ROW_BG),
            ),
        }

    @staticmethod
    def total_row():
        return {
            "font": Font(name="Arial", size=10, bold=True, color=ExportColors.NORMAL_TEXT),
            "fill": PatternFill("solid", fgColor=ExportColors.TOTAL_ROW_BG),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                top=Side(style="medium", color=ExportColors.SECTION_BORDER),
                bottom=Side(style="medium", color=ExportColors.SECTION_BORDER),
            ),
        }


class ExcelExporter:
    """Reusable workbook builder with styling helpers."""

    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def create_cover_sheet(self, metadata: Dict[str, str]) -> None:
        ws = self.workbook.create_sheet("Cover", 0)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        row = 1

        ws.merge_cells(f"A{row}:B{row+1}")
        cell = ws[f"A{row}"]
        cell.value = metadata.get("report_title", "GLOBAL REPORT")
        styles = ExportStyles.cover_title()
        cell.font = styles["font"]
        cell.fill = styles["fill"]
        cell.alignment = styles["alignment"]
        ws.row_dimensions[row].height = 30
        ws.row_dimensions[row + 1].height = 30
        row += 3

        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = metadata.get("report_subtitle", "Data Intelligence Workbook")
        styles = ExportStyles.cover_subtitle()
        cell.font = styles["font"]
        cell.fill = styles["fill"]
        cell.alignment = styles["alignment"]
        ws.row_dimensions[row].height = 24
        row += 2

        self._section_header(ws, row, "REPORT METADATA")
        row += 2
        info = [
            ("Generated", metadata.get("generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))),
            ("Coverage", f"{metadata.get('start_date', 'N/A')} → {metadata.get('end_date', 'N/A')}"),
            ("Scope", metadata.get("scope", "Global")),
        ]
        for label, value in info:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        row += 1

        self._section_header(ws, row, "KEY METRICS")
        row += 2
        for key, value in metadata.get("key_metrics", {}).items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
            row += 1

    def add_data_sheet(self, title: str, sections: List[Dict]) -> None:
        ws = self.workbook.create_sheet(title[:31])
        current_row = 1

        ws.merge_cells(f"A{current_row}:E{current_row}")
        cell = ws[f"A{current_row}"]
        cell.value = title
        cell.font = Font(name="Arial", size=14, bold=True, color=ExportColors.HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=ExportColors.SUBHEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 2

        for section in sections:
            df = section.get("data", pd.DataFrame())
            chart_type = section.get("chart_type", "table")
            description = section.get("description", "")
            df_raw = section.get("data_raw")
            chart_cfg = section.get("chart_config", {})
            current_row = self._write_section(
                ws,
                current_row,
                section.get("title", "Untitled"),
                df,
                chart_type,
                description,
                df_raw,
                chart_cfg,
            )
            current_row += 2

    def add_data_dictionary_sheet(self) -> None:
        ws = self.workbook.create_sheet("Data Dictionary")
        entries = pd.DataFrame(
            {
                "Metric": [
                    "OBS_VALUE",
                    "TIME_PERIOD",
                    "REF_AREA",
                    "REF_AREA_LABEL",
                    "region",
                    "era",
                    "decade",
                    "year",
                    "month",
                ],
                "Description": [
                    "Food price inflation value",
                    "Monthly period (date)",
                    "Country/area ISO code",
                    "Country/area name",
                    "Assigned region grouping",
                    "Structural era classification",
                    "Decade bucket (YYYY)",
                    "Calendar year",
                    "Calendar month",
                ],
            }
        )
        self._write_section(ws, 1, "Dataset Columns", entries, "table", "")

    def save_to_bytes(self) -> bytes:
        buff = io.BytesIO()
        self.workbook.save(buff)
        buff.seek(0)
        return buff.getvalue()

    # ----- internal helpers -----
    def _section_header(self, ws, row: int, text: str) -> None:
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        styles = ExportStyles.section_header()
        cell.value = text
        cell.font = styles["font"]
        cell.fill = styles["fill"]
        cell.alignment = styles["alignment"]

    def _write_section(
        self,
        ws,
        row: int,
        title: str,
        df: pd.DataFrame,
        chart_type: str,
        description: str,
        df_raw: Optional[pd.DataFrame] = None,
        chart_cfg: Optional[Dict] = None,
    ) -> int:
        col_count = max(1, len(df.columns))
        ws.merge_cells(f"A{row}:{get_column_letter(col_count)}{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(name="Arial", size=12, bold=True, color=ExportColors.NORMAL_TEXT)
        cell.fill = PatternFill("solid", fgColor=ExportColors.ALT_ROW_BG)
        row += 1

        if description:
            ws.merge_cells(f"A{row}:{get_column_letter(col_count)}{row}")
            desc_cell = ws[f"A{row}"]
            desc_cell.value = description
            desc_cell.font = Font(name="Arial", size=9, italic=True, color=ExportColors.MUTED_TEXT)
            row += 1

        if df.empty:
            ws[f"A{row}"] = "No data available."
            return row + 1

        header_row = row
        for idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=idx)
            cell.value = column
            styles = ExportStyles.table_header()
            cell.font = styles["font"]
            cell.fill = styles["fill"]
            cell.alignment = styles["alignment"]
            cell.border = styles["border"]
        row += 1

        for r_index, (_, values) in enumerate(df.iterrows(), start=row):
            is_alt = (r_index - row) % 2 == 1
            for c_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=r_index, column=c_idx)
                cell.value = self._format_cell_value(value)
                styles = ExportStyles.data_cell(is_alt)
                cell.font = styles["font"]
                cell.fill = styles["fill"]
                cell.alignment = styles["alignment"]
                cell.border = styles["border"]
        row = row + len(df)
        self._auto_width(ws, df)

        if chart_type in {"bar", "line"} and df_raw is not None and not df_raw.empty:
            row += 1
            self._insert_chart(ws, row, df_raw, chart_type, chart_cfg)
            row += 15

        return row

    @staticmethod
    def _auto_width(ws, df: pd.DataFrame) -> None:
        for idx, col in enumerate(df.columns, start=1):
            max_len = max([len(str(col))] + [len(str(val)) for val in df[col]])
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)

    def _insert_chart(self, ws, row: int, df_raw: pd.DataFrame, chart_type: str, cfg: Dict) -> None:
        data_col = cfg.get("data_col")
        category_col = cfg.get("category_col")
        if not data_col or not category_col:
            return

        start_col = 30
        for idx, value in enumerate(df_raw[category_col], start=row):
            cell = ws.cell(row=idx, column=start_col)
            cell.value = self._format_cell_value(value)
            cell.font = Font(color="FFFFFF")
        for idx, value in enumerate(df_raw[data_col], start=row):
            cell = ws.cell(row=idx, column=start_col + 1)
            cell.value = value
            cell.font = Font(color="FFFFFF")

        cats = Reference(ws, min_col=start_col, min_row=row, max_row=row + len(df_raw) - 1)
        vals = Reference(ws, min_col=start_col + 1, min_row=row, max_row=row + len(df_raw) - 1)

        if chart_type == "bar":
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
        else:
            chart = LineChart()
            chart.style = 12

        chart.title = cfg.get("title", "")
        chart.add_data(vals, titles_from_data=False)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showValue = True
        chart.height = 12
        chart.width = 24
        ws.add_chart(chart, f"E{row}")

    @staticmethod
    def _format_cell_value(value):
        if isinstance(value, pd.Interval):
            left = value.left
            right = value.right
            def fmt(val):
                if isinstance(val, (int, float)):
                    return f"{val:.2f}"
                return str(val)
            return f"{fmt(left)} → {fmt(right)}"
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value


def format_number(num: float) -> str:
    if pd.isna(num):
        return ""
    if num >= 1_000_000:
        val = num / 1_000_000
        return f"{val:.2f}M".rstrip("0").rstrip(".")
    if num >= 1_000:
        val = num / 1_000
        return f"{val:.2f}K".rstrip("0").rstrip(".")
    return f"{num:.0f}"


def format_currency(num: float) -> str:
    if pd.isna(num):
        return ""
    if num >= 1_000_000:
        val = num / 1_000_000
        return f"${val:.2f}M".rstrip("0").rstrip(".")
    if num >= 1_000:
        val = num / 1_000
        return f"${val:.2f}K".rstrip("0").rstrip(".")
    return f"${num:.2f}"


def format_percentage(num: float) -> str:
    if pd.isna(num):
        return ""
    return f"{num:.2f}%"
