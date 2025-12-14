"""
Excel Export Module for Merch Intelligence Hub
Exports dashboard data to formatted Excel workbooks with multiple sheets
"""

import pandas as pd
import io
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


# ============================================================================
# COLOR SCHEME CONFIGURATION (Customizable)
# ============================================================================
class ExportColors:
    """Centralized color scheme for Excel exports - easy to customize"""

    # Primary Colors (matching dashboard theme)
    PRIMARY_BLUE = "1F77B4"      # Main header background
    PRIMARY_GREEN = "2CA02C"     # Positive growth
    PRIMARY_RED = "D62728"       # Negative growth
    PRIMARY_ORANGE = "FF7F0E"    # Warning/attention

    # Background Colors
    HEADER_BG = "2C3E50"         # Dark blue-gray for main headers
    SUBHEADER_BG = "34495E"      # Slightly lighter for sub-headers
    TABLE_HEADER_BG = "5DADE2"   # Light blue for table headers
    ALT_ROW_BG = "F8F9F9"        # Light gray for alternating rows
    TOTAL_ROW_BG = "D5DBDB"      # Gray for total/summary rows

    # Text Colors
    HEADER_TEXT = "FFFFFF"       # White text for dark backgrounds
    NORMAL_TEXT = "000000"       # Black text for light backgrounds
    MUTED_TEXT = "7F8C8D"        # Gray for less important text

    # Accent Colors
    COVER_BG = "1A5490"          # Cover page background
    SECTION_BORDER = "34495E"    # Section divider borders

    # Conditional Formatting
    POSITIVE_BG = "D5F4E6"       # Light green background
    NEGATIVE_BG = "FADBD8"       # Light red background
    NEUTRAL_BG = "F8F9F9"        # Light gray background


class ExportStyles:
    """Pre-defined cell styles for consistent formatting"""

    @staticmethod
    def get_cover_title_style():
        """Style for cover page main title"""
        return {
            'font': Font(name='Arial', size=24, bold=True, color=ExportColors.HEADER_TEXT),
            'fill': PatternFill(start_color=ExportColors.COVER_BG, end_color=ExportColors.COVER_BG, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

    @staticmethod
    def get_cover_subtitle_style():
        """Style for cover page subtitle"""
        return {
            'font': Font(name='Arial', size=14, italic=True, color=ExportColors.HEADER_TEXT),
            'fill': PatternFill(start_color=ExportColors.COVER_BG, end_color=ExportColors.COVER_BG, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

    @staticmethod
    def get_section_header_style():
        """Style for section headers (e.g., REPORT METADATA)"""
        return {
            'font': Font(name='Arial', size=12, bold=True, color=ExportColors.HEADER_TEXT),
            'fill': PatternFill(start_color=ExportColors.HEADER_BG, end_color=ExportColors.HEADER_BG, fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center')
        }

    @staticmethod
    def get_table_header_style():
        """Style for table column headers"""
        return {
            'font': Font(name='Arial', size=11, bold=True, color=ExportColors.HEADER_TEXT),
            'fill': PatternFill(start_color=ExportColors.TABLE_HEADER_BG, end_color=ExportColors.TABLE_HEADER_BG, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color=ExportColors.SECTION_BORDER),
                right=Side(style='thin', color=ExportColors.SECTION_BORDER),
                top=Side(style='thin', color=ExportColors.SECTION_BORDER),
                bottom=Side(style='thin', color=ExportColors.SECTION_BORDER)
            )
        }

    @staticmethod
    def get_data_cell_style(is_alt_row=False):
        """Style for regular data cells"""
        bg_color = ExportColors.ALT_ROW_BG if is_alt_row else ExportColors.HEADER_TEXT
        return {
            'font': Font(name='Arial', size=10, color=ExportColors.NORMAL_TEXT),
            'fill': PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin', color=ExportColors.ALT_ROW_BG),
                right=Side(style='thin', color=ExportColors.ALT_ROW_BG),
                top=Side(style='thin', color=ExportColors.ALT_ROW_BG),
                bottom=Side(style='thin', color=ExportColors.ALT_ROW_BG)
            )
        }

    @staticmethod
    def get_total_row_style():
        """Style for total/summary rows"""
        return {
            'font': Font(name='Arial', size=10, bold=True, color=ExportColors.NORMAL_TEXT),
            'fill': PatternFill(start_color=ExportColors.TOTAL_ROW_BG, end_color=ExportColors.TOTAL_ROW_BG, fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                top=Side(style='medium', color=ExportColors.SECTION_BORDER),
                bottom=Side(style='medium', color=ExportColors.SECTION_BORDER)
            )
        }

    @staticmethod
    def get_growth_cell_style(value):
        """Style for growth cells with conditional formatting"""
        if pd.isna(value):
            bg_color = ExportColors.NEUTRAL_BG
            font_color = ExportColors.MUTED_TEXT
        elif value > 0:
            bg_color = ExportColors.POSITIVE_BG
            font_color = ExportColors.PRIMARY_GREEN
        elif value < 0:
            bg_color = ExportColors.NEGATIVE_BG
            font_color = ExportColors.PRIMARY_RED
        else:
            bg_color = ExportColors.NEUTRAL_BG
            font_color = ExportColors.NORMAL_TEXT

        return {
            'font': Font(name='Arial', size=10, bold=True, color=font_color),
            'fill': PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }


class ExcelExporter:
    """Main Excel export engine"""

    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

    def create_cover_sheet(self, metadata: Dict) -> None:
        """Create formatted cover sheet with report metadata"""
        ws = self.workbook.create_sheet("Cover", 0)

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

        current_row = 1

        # Main title
        ws.merge_cells(f'A{current_row}:B{current_row + 1}')
        cell = ws[f'A{current_row}']
        cell.value = "MERCH INTELLIGENCE HUB"
        styles = ExportStyles.get_cover_title_style()
        cell.font = styles['font']
        cell.fill = styles['fill']
        cell.alignment = styles['alignment']
        ws.row_dimensions[current_row].height = 30
        ws.row_dimensions[current_row + 1].height = 30
        current_row += 3

        # Subtitle
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "Comprehensive Analytics Report"
        styles = ExportStyles.get_cover_subtitle_style()
        cell.font = styles['font']
        cell.fill = styles['fill']
        cell.alignment = styles['alignment']
        ws.row_dimensions[current_row].height = 25
        current_row += 3

        # Section: Report Metadata
        self._add_section_header(ws, current_row, "REPORT METADATA")
        current_row += 1
        self._add_separator_line(ws, current_row)
        current_row += 1

        ws[f'A{current_row}'] = "Generated:"
        ws[f'B{current_row}'] = metadata.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        current_row += 1

        ws[f'A{current_row}'] = "Report Period:"
        ws[f'B{current_row}'] = f"{metadata.get('start_date', 'N/A')} to {metadata.get('end_date', 'N/A')}"
        current_row += 1

        ws[f'A{current_row}'] = "Comparison Mode:"
        ws[f'B{current_row}'] = metadata.get('comparison_mode', 'N/A')
        current_row += 2

        # Section: Filters Applied
        self._add_section_header(ws, current_row, "FILTERS APPLIED")
        current_row += 1
        self._add_separator_line(ws, current_row)
        current_row += 1

        filters = metadata.get('filters', {})
        for filter_name, filter_value in filters.items():
            ws[f'A{current_row}'] = f"{filter_name}:"
            ws[f'B{current_row}'] = filter_value
            current_row += 1

        current_row += 1

        # Section: Key Metrics Summary
        self._add_section_header(ws, current_row, "KEY METRICS SUMMARY")
        current_row += 1
        self._add_separator_line(ws, current_row)
        current_row += 1

        metrics = metadata.get('key_metrics', {})
        for metric_name, metric_data in metrics.items():
            ws[f'A{current_row}'] = f"{metric_name}:"
            ws[f'B{current_row}'] = metric_data
            current_row += 1

        current_row += 1

        # Section: Sheets in Report
        self._add_section_header(ws, current_row, "SHEETS IN THIS REPORT")
        current_row += 1
        self._add_separator_line(ws, current_row)
        current_row += 1

        sheets = metadata.get('sheets', [])
        for idx, sheet_name in enumerate(sheets, 1):
            ws[f'A{current_row}'] = f"{idx}."
            ws[f'B{current_row}'] = sheet_name
            current_row += 1

    def _add_section_header(self, ws, row: int, text: str) -> None:
        """Add formatted section header"""
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = text
        styles = ExportStyles.get_section_header_style()
        cell.font = styles['font']
        cell.fill = styles['fill']
        cell.alignment = styles['alignment']
        ws.row_dimensions[row].height = 20

    def _add_separator_line(self, ws, row: int) -> None:
        """Add separator line under section headers"""
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "─" * 80
        cell.font = Font(color=ExportColors.SECTION_BORDER)

    def add_data_sheet(self, sheet_name: str, sections: List[Dict]) -> None:
        """
        Add a data sheet with multiple chart/table sections

        Args:
            sheet_name: Name of the sheet
            sections: List of dictionaries with structure:
                {
                    'title': 'Chart/Table Title',
                    'data': pd.DataFrame,
                    'data_raw': pd.DataFrame (optional - numeric data for charts),
                    'chart_type': 'bar', 'line', 'pie', or 'table',
                    'chart_config': {...} (optional - chart configuration),
                    'description': 'Optional description'
                }
        """
        ws = self.workbook.create_sheet(sheet_name)
        current_row = 1

        # Sheet title
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = sheet_name
        styles = ExportStyles.get_section_header_style()
        cell.font = Font(name='Arial', size=14, bold=True, color=ExportColors.HEADER_TEXT)
        cell.fill = styles['fill']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 25
        current_row += 2

        # Add each section
        for section in sections:
            current_row = self._add_table_section(
                ws,
                current_row,
                section.get('title', 'Untitled'),
                section.get('data', pd.DataFrame()),
                section.get('chart_type', 'table'),
                section.get('description', ''),
                section.get('data_raw', None),
                section.get('chart_config', {})
            )
            current_row += 2  # Space between sections

    def _add_table_section(self, ws, start_row: int, title: str, df: pd.DataFrame,
                           chart_type: str, description: str, df_raw: pd.DataFrame = None,
                           chart_config: dict = None) -> int:
        """
        Add a formatted table section with optional chart

        Args:
            ws: Worksheet
            start_row: Starting row
            title: Section title
            df: Formatted data (with K/M notation)
            chart_type: 'bar', 'line', 'pie', or 'table'
            description: Section description
            df_raw: Raw numeric data for charts (optional)
            chart_config: Chart configuration dict (optional)
        """
        current_row = start_row
        if chart_config is None:
            chart_config = {}

        # Section title
        ws.merge_cells(f'A{current_row}:{get_column_letter(len(df.columns))}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = title
        cell.font = Font(name='Arial', size=12, bold=True, color=ExportColors.NORMAL_TEXT)
        cell.fill = PatternFill(start_color=ExportColors.SUBHEADER_BG,
                               end_color=ExportColors.SUBHEADER_BG, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Description (if provided)
        if description:
            ws.merge_cells(f'A{current_row}:{get_column_letter(len(df.columns))}{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = description
            cell.font = Font(name='Arial', size=9, italic=True, color=ExportColors.MUTED_TEXT)
            current_row += 1

        # Add table
        if not df.empty:
            # Headers
            header_row = current_row
            for col_idx, column in enumerate(df.columns, 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.value = column
                styles = ExportStyles.get_table_header_style()
                cell.font = styles['font']
                cell.fill = styles['fill']
                cell.alignment = styles['alignment']
                cell.border = styles['border']

            ws.row_dimensions[header_row].height = 18
            current_row += 1

            # Data rows
            for row_idx, row_data in enumerate(df.itertuples(index=False), start=current_row):
                is_alt_row = (row_idx - current_row) % 2 == 1
                is_total_row = False

                # Check if this is a total row (looks for "TOTAL", "Others", etc.)
                first_value = str(row_data[0]) if row_data else ""
                if any(keyword in first_value.upper() for keyword in ['TOTAL', 'OTHERS', 'SUM', 'AVERAGE']):
                    is_total_row = True

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value

                    # Apply appropriate style
                    if is_total_row:
                        styles = ExportStyles.get_total_row_style()
                        cell.font = styles['font']
                        cell.fill = styles['fill']
                        cell.alignment = styles['alignment']
                        cell.border = styles['border']
                    else:
                        styles = ExportStyles.get_data_cell_style(is_alt_row)
                        cell.font = styles['font']
                        cell.fill = styles['fill']
                        cell.alignment = styles['alignment']
                        cell.border = styles['border']

                    # Apply number formatting
                    column_name = df.columns[col_idx - 1]
                    cell.number_format = self._get_number_format(column_name, value)

                    # Apply conditional formatting for growth columns
                    if any(keyword in column_name.lower() for keyword in ['growth', 'change', 'δ', 'Δ']):
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            growth_styles = ExportStyles.get_growth_cell_style(value)
                            cell.font = growth_styles['font']
                            cell.fill = growth_styles['fill']
                            cell.alignment = growth_styles['alignment']

                current_row = row_idx

            current_row += 1

            # Auto-adjust column widths
            self._auto_adjust_columns(ws, df)

            # Add chart if configured and raw data provided
            if df_raw is not None and chart_type in ['bar', 'line', 'pie']:
                chart_row = current_row + 1

                # Get configuration
                data_col = chart_config.get('data_col', None)
                category_col = chart_config.get('category_col', df_raw.columns[0])
                chart_title = chart_config.get('title', title)

                if data_col and data_col in df_raw.columns:
                    # Write raw data to hidden columns (far right of spreadsheet)
                    hidden_start_col = 50  # Column AX - far enough to not interfere

                    # Write category column with white font (invisible but not hidden)
                    row_offset = header_row + 1
                    for idx, value in enumerate(df_raw[category_col]):
                        cell = ws.cell(row=row_offset + idx, column=hidden_start_col, value=value)
                        cell.font = Font(color="FFFFFF")  # White font color

                    # Write data column with white font (invisible but not hidden)
                    for idx, value in enumerate(df_raw[data_col]):
                        cell = ws.cell(row=row_offset + idx, column=hidden_start_col + 1, value=value)
                        cell.font = Font(color="FFFFFF")  # White font color

                    # Set column width to minimal (but don't hide) so columns are not visible
                    ws.column_dimensions[get_column_letter(hidden_start_col)].width = 0.5
                    ws.column_dimensions[get_column_letter(hidden_start_col + 1)].width = 0.5

                    # Calculate number of data rows
                    num_rows = len(df_raw)
                    data_start_row = header_row + 1
                    data_end_row = data_start_row + num_rows - 1

                    # Build ranges using hidden columns
                    cat_letter = get_column_letter(hidden_start_col)
                    data_letter = get_column_letter(hidden_start_col + 1)

                    categories_range = f"{cat_letter}{data_start_row}:{cat_letter}{data_end_row}"
                    data_range = f"{data_letter}{data_start_row}:{data_letter}{data_end_row}"

                    # Position chart at column E (or after table if wider)
                    chart_col = max(5, len(df.columns) + 2)
                    chart_position = f"{get_column_letter(chart_col)}{chart_row}"

                    # Add appropriate chart type
                    if chart_type == 'bar':
                        self.add_bar_chart(ws, data_range, categories_range,
                                         chart_title, chart_position)
                    elif chart_type == 'line':
                        self.add_line_chart(ws, data_range, categories_range,
                                          chart_title, chart_position)
                    elif chart_type == 'pie':
                        self.add_pie_chart(ws, data_range, categories_range,
                                         chart_title, chart_position)

                    # Update current row to account for chart height (~15 rows)
                    current_row = chart_row + 15

        return current_row

    def _get_number_format(self, column_name: str, value) -> str:
        """Get appropriate number format based on column name and value"""
        col_lower = column_name.lower()

        # Currency - values are already formatted as strings with K/M suffix
        if any(keyword in col_lower for keyword in ['revenue', 'price', 'aov', 'value', '$']):
            return '@'  # Text format to preserve K/M formatting

        # Percentage
        if any(keyword in col_lower for keyword in ['%', 'rate', 'cr', 'ctr', 'share']):
            if isinstance(value, str):
                return '@'  # Already formatted as string with %
            elif isinstance(value, (int, float)) and abs(value) < 1:
                return '0.00%'
            else:
                return '0.00"%"'  # Already in percentage form

        # Growth columns - preserve +/- and % formatting
        if any(keyword in col_lower for keyword in ['growth', 'change', 'δ', 'Δ']):
            return '@'  # Text format to preserve +15.3% formatting

        # Large numbers (sessions, views, etc.) - also formatted with K/M
        if any(keyword in col_lower for keyword in ['sessions', 'views', 'clicks', 'users', 'purchases', 'items', 'count']):
            return '@'  # Text format to preserve K/M formatting

        # Decimal numbers
        if isinstance(value, float):
            return '#,##0.00'

        # Default
        return 'General'

    def _auto_adjust_columns(self, ws, df: pd.DataFrame) -> None:
        """Auto-adjust column widths based on content"""
        for idx, column in enumerate(df.columns, 1):
            max_length = len(str(column))
            for value in df[column]:
                try:
                    value_length = len(str(value))
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    def add_bar_chart(self, ws, data_range: str, categories_range: str,
                      title: str, position: str, height: int = 15, width: int = 25) -> None:
        """
        Add a bar chart to the worksheet

        Args:
            ws: Worksheet object
            data_range: Range for data values (e.g., 'B2:B11')
            categories_range: Range for category labels (e.g., 'A2:A11')
            title: Chart title
            position: Cell position for chart (e.g., 'E2')
            height: Chart height in rows
            width: Chart width in columns
        """
        from openpyxl.utils import column_index_from_string, coordinate_to_tuple

        chart = BarChart()
        chart.type = "col"  # Column chart
        chart.style = 11  # Modern blue style
        chart.title = title
        chart.y_axis.title = "Value"
        chart.y_axis.majorGridlines = None  # Cleaner look

        # Parse data range (e.g., 'B2:B11' -> min_col=2, min_row=2, max_col=2, max_row=11)
        data_start, data_end = data_range.split(':')
        data_start_row, data_start_col = coordinate_to_tuple(data_start)
        data_end_row, data_end_col = coordinate_to_tuple(data_end)

        cats_start, cats_end = categories_range.split(':')
        cats_start_row, cats_start_col = coordinate_to_tuple(cats_start)
        cats_end_row, cats_end_col = coordinate_to_tuple(cats_end)

        data = Reference(ws, min_col=data_start_col, min_row=data_start_row,
                        max_col=data_end_col, max_row=data_end_row)
        cats = Reference(ws, min_col=cats_start_col, min_row=cats_start_row,
                        max_col=cats_end_col, max_row=cats_end_row)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        chart.height = height
        chart.width = width

        ws.add_chart(chart, position)

    def add_line_chart(self, ws, data_range: str, categories_range: str,
                       title: str, position: str, height: int = 15, width: int = 25) -> None:
        """Add a line chart to the worksheet"""
        from openpyxl.utils import coordinate_to_tuple

        chart = LineChart()
        chart.style = 12  # Modern style
        chart.title = title
        chart.y_axis.title = "Value"
        chart.y_axis.majorGridlines = None  # Cleaner look
        chart.smooth = True  # Smooth lines

        # Parse ranges
        data_start, data_end = data_range.split(':')
        data_start_row, data_start_col = coordinate_to_tuple(data_start)
        data_end_row, data_end_col = coordinate_to_tuple(data_end)

        cats_start, cats_end = categories_range.split(':')
        cats_start_row, cats_start_col = coordinate_to_tuple(cats_start)
        cats_end_row, cats_end_col = coordinate_to_tuple(cats_end)

        data = Reference(ws, min_col=data_start_col, min_row=data_start_row,
                        max_col=data_end_col, max_row=data_end_row)
        cats = Reference(ws, min_col=cats_start_col, min_row=cats_start_row,
                        max_col=cats_end_col, max_row=cats_end_row)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        chart.height = height
        chart.width = width

        ws.add_chart(chart, position)

    def add_pie_chart(self, ws, data_range: str, categories_range: str,
                      title: str, position: str, height: int = 15, width: int = 15) -> None:
        """Add a pie chart to the worksheet"""
        from openpyxl.utils import coordinate_to_tuple

        chart = PieChart()
        chart.title = title
        chart.style = 26  # Modern style with 3D effect

        # Parse ranges
        data_start, data_end = data_range.split(':')
        data_start_row, data_start_col = coordinate_to_tuple(data_start)
        data_end_row, data_end_col = coordinate_to_tuple(data_end)

        cats_start, cats_end = categories_range.split(':')
        cats_start_row, cats_start_col = coordinate_to_tuple(cats_start)
        cats_end_row, cats_end_col = coordinate_to_tuple(cats_end)

        data = Reference(ws, min_col=data_start_col, min_row=data_start_row,
                        max_col=data_end_col, max_row=data_end_row)
        cats = Reference(ws, min_col=cats_start_col, min_row=cats_start_row,
                        max_col=cats_end_col, max_row=cats_end_row)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # Add data labels showing percentages
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True

        chart.height = height
        chart.width = width

        ws.add_chart(chart, position)

    def add_data_dictionary_sheet(self) -> None:
        """Add data dictionary sheet explaining all metrics and columns"""
        ws = self.workbook.create_sheet("Data Dictionary")

        # Dictionary data
        dictionary_data = {
            'Metric/Column': [
                'Item Revenue', 'Items Purchased', 'Items Viewed', 'Items Added to Cart',
                'Items Clicked in List', 'Items Viewed in List', 'Sessions', 'Total Users',
                'Conversion Rate (CR)', 'Click-Through Rate (CTR)', 'Cart Add Rate',
                'Average Order Value (AOV)', 'Revenue per Session', 'Cart Abandonment Rate',
                'Engagement Score', 'Growth %', 'Share %', 'Δ% (Delta %)',
                'List Type', 'SMR Type', 'Session Nahdi channel grouping'
            ],
            'Description': [
                'Total revenue generated from item purchases',
                'Number of items purchased',
                'Number of times items were viewed in detail',
                'Number of items added to shopping cart',
                'Number of times items were clicked from a list',
                'Number of times items appeared in a list',
                'Number of user sessions',
                'Total unique users',
                'Percentage of viewed items that were purchased',
                'Percentage of list views that resulted in clicks',
                'Percentage of viewed items that were added to cart',
                'Average revenue per purchased item',
                'Average revenue generated per session',
                'Percentage of cart additions that did not result in purchase',
                'Composite metric combining views, clicks, and cart additions',
                'Percentage change compared to previous period',
                'Percentage of total for this dimension',
                'Change in share percentage (in percentage points)',
                'Classification of item list (DY PDP, plp, clp, search, etc.)',
                'Strategic classification (Recommendation, Merch, Search, PDP)',
                'Traffic source channel grouping'
            ],
            'Formula/Logic': [
                'Sum of item_revenue', 'Sum of items_purchased', 'Sum of items_viewed',
                'Sum of items_added_to_cart', 'Sum of items_clicked_in_list',
                'Sum of items_viewed_in_list', 'Sum of sessions', 'Sum of total_users',
                '(Items Purchased / Items Viewed) × 100',
                '(Items Clicked in List / Items Viewed in List) × 100',
                '(Items Added to Cart / Items Viewed) × 100',
                'Item Revenue / Items Purchased',
                'Item Revenue / Sessions',
                '((Items Added to Cart - Items Purchased) / Items Added to Cart) × 100',
                '(views + clicks×2 + cart adds×3) / sessions',
                '((Current - Previous) / Previous) × 100',
                '(Value / Total) × 100',
                'Current Share % - Previous Share %',
                'Based on item_list_id and Contentful metadata',
                'Derived from List Type classification',
                'GA4 traffic source dimension'
            ]
        }

        df_dict = pd.DataFrame(dictionary_data)
        sections = [{
            'title': 'Metrics and Columns Reference',
            'data': df_dict,
            'chart_type': 'table',
            'description': 'Definitions and calculation methods for all metrics used in this report'
        }]

        self._add_table_section(ws, 1, 'Data Dictionary', df_dict, 'table',
                               'Reference guide for understanding report metrics and columns')

    def save_to_bytes(self) -> bytes:
        """Save workbook to bytes for download"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()


# ============================================================================
# Helper Functions for Number Formatting (K/M format matching dashboard)
# ============================================================================

def format_number(num):
    """Format numbers with K, M suffixes (matching dashboard format)"""
    if pd.isna(num):
        return ""
    if num >= 1_000_000:
        val = num / 1_000_000
        # Show decimals only if needed (not whole number)
        if val == int(val):
            return f"{int(val)}M"
        else:
            formatted = f"{val:.2f}".rstrip('0').rstrip('.')
            return f"{formatted}M"
    elif num >= 1_000:
        val = num / 1_000
        # Show decimals only if needed (not whole number)
        if val == int(val):
            return f"{int(val)}K"
        else:
            formatted = f"{val:.2f}".rstrip('0').rstrip('.')
            return f"{formatted}K"
    else:
        return f"{int(num)}" if num == int(num) else f"{num:.0f}"


def format_currency(num):
    """Format currency with K, M suffixes (matching dashboard format)"""
    if pd.isna(num):
        return ""
    if num >= 1_000_000:
        val = num / 1_000_000
        # Show decimals only if needed (not whole number)
        if val == int(val):
            return f"${int(val)}M"
        else:
            formatted = f"{val:.2f}".rstrip('0').rstrip('.')
            return f"${formatted}M"
    elif num >= 1_000:
        val = num / 1_000
        # Show decimals only if needed (not whole number)
        if val == int(val):
            return f"${int(val)}K"
        else:
            formatted = f"{val:.2f}".rstrip('0').rstrip('.')
            return f"${formatted}K"
    else:
        return f"${num:.2f}"


def format_percentage(num):
    """Format percentage with 2 decimals"""
    if pd.isna(num):
        return ""
    return f"{num:.2f}%"


def format_growth(num):
    """Format growth with +/- prefix and %"""
    if pd.isna(num):
        return ""
    if num > 0:
        return f"+{num:.1f}%"
    else:
        return f"{num:.1f}%"


# ============================================================================
# Helper Functions for Data Preparation
# ============================================================================

def prepare_top_n_with_others(df: pd.DataFrame, value_col: str, name_col: str,
                              top_n: int = 20) -> pd.DataFrame:
    """
    Prepare dataframe with top N rows and 'Others' row for remaining

    Args:
        df: Source dataframe
        value_col: Column to sort by (e.g., 'Revenue')
        name_col: Column with names (e.g., 'Category')
        top_n: Number of top items to keep

    Returns:
        DataFrame with top N + Others row
    """
    if len(df) <= top_n:
        return df.copy()

    # Sort by value column
    df_sorted = df.sort_values(value_col, ascending=False).reset_index(drop=True)

    # Get top N
    top_df = df_sorted.head(top_n).copy()

    # Calculate Others row
    others_df = df_sorted.iloc[top_n:].copy()
    others_row = {}
    others_row[name_col] = 'Others'

    for col in df.columns:
        if col == name_col:
            continue
        elif df[col].dtype in ['float64', 'int64']:
            others_row[col] = others_df[col].sum()
        else:
            others_row[col] = ''

    # Combine
    result_df = pd.concat([top_df, pd.DataFrame([others_row])], ignore_index=True)

    return result_df


def add_total_row(df: pd.DataFrame, name_col: str) -> pd.DataFrame:
    """Add TOTAL row to dataframe"""
    total_row = {}
    total_row[name_col] = 'TOTAL'

    for col in df.columns:
        if col == name_col:
            continue
        elif df[col].dtype in ['float64', 'int64']:
            total_row[col] = df[col].sum()
        else:
            total_row[col] = ''

    result_df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    return result_df


def format_growth_column(df: pd.DataFrame, col_name: str) -> pd.DataFrame:
    """Format growth column with + prefix and % suffix"""
    df = df.copy()
    if col_name in df.columns:
        df[col_name] = df[col_name].apply(lambda x: f"+{x:.1f}%" if x > 0 else f"{x:.1f}%" if pd.notna(x) else "")
    return df


def add_trend_arrows(df: pd.DataFrame, growth_col: str, new_col: str = 'Trend') -> pd.DataFrame:
    """Add trend arrow column based on growth values"""
    df = df.copy()

    def get_arrow(value):
        if pd.isna(value):
            return "→"
        elif value > 0:
            return "↑"
        elif value < 0:
            return "↓"
        else:
            return "→"

    df[new_col] = df[growth_col].apply(get_arrow)
    return df
